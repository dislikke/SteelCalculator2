from __future__ import annotations

from io import BytesIO
from pathlib import Path
from copy import copy

from openpyxl import load_workbook

from .optimizer import default_params

BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = BASE_DIR / "resources" / "excel_template.xlsx"

EPS = 1e-9

def to_float(value, default=0.0):
    if value is None:
        return default
    return float(value)

def ge_with_tol(actual, required, eps=EPS):
    return to_float(actual) + eps >= to_float(required)

def le_with_tol(actual, limit, eps=EPS):
    return to_float(actual) <= to_float(limit) + eps

def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def safe_coef(coef: dict | None) -> dict:
    defaults = default_params()["coef"]
    coef = coef or {}
    return {
        "sigma_base": float(coef.get("sigma_base", defaults["sigma_base"])),
        "sigma_Cr": float(coef.get("sigma_Cr", defaults["sigma_Cr"])),
        "sigma_Mo": float(coef.get("sigma_Mo", defaults["sigma_Mo"])),
        "sigma_CrMo": float(coef.get("sigma_CrMo", defaults["sigma_CrMo"])),
        "hrc_base": float(coef.get("hrc_base", defaults["hrc_base"])),
        "hrc_Ni": float(coef.get("hrc_Ni", defaults["hrc_Ni"])),
        "hrc_Mn": float(coef.get("hrc_Mn", defaults["hrc_Mn"])),
        "hrc_NiMn": float(coef.get("hrc_NiMn", defaults["hrc_NiMn"])),
        "T_base": float(coef.get("T_base", defaults["T_base"])),
        "T_drop": float(coef.get("T_drop", defaults["T_drop"])),
    }


def result_req(result) -> dict:
    if result.variant:
        return {
            "sigma": result.variant.sigma_req,
            "hrc": result.variant.hard_req,
            "t": result.variant.t_req,
        }
    return result.req or {"sigma": 0.0, "hrc": 0.0, "t": 0.0}


def result_limits(result) -> dict:
    if result.variant:
        return {
            "sum_min": result.variant.sum_min,
            "sum_max": result.variant.sum_max,
            "crni_max": result.variant.crni_max,
        }
    return result.limits or {"sum_min": 0.0, "sum_max": 6.0, "crni_max": 2.0}


def build_excel_export(variants, results) -> BytesIO:
    wb = load_workbook(TEMPLATE_PATH)

    ws_variants = wb["Варианты"]
    ws_results = wb["Результаты"]

    start_row = 3
    max_col_variants = ws_variants.max_column
    max_col_results = ws_results.max_column

    variant_key_map = {}
    used_keys = set()

    # ---- ЛИСТ ВАРИАНТЫ ----
    current_row = start_row

    for idx, v in enumerate(variants, start=1):
        if current_row > start_row:
            copy_row_style(ws_variants, start_row, current_row, max_col_variants)

        excel_key = f"V{idx:03d}"
        while excel_key in used_keys:
            idx += 1
            excel_key = f"V{idx:03d}"

        used_keys.add(excel_key)
        variant_key_map[v.id] = excel_key

        coef = safe_coef(v.coef)

        row = [
            excel_key,                  # excel_key
            v.name,                     # name
            "db",                       # source_type
            v.id,                       # source_variant_id
            v.cr_min,
            v.cr_max,
            v.ni_min,
            v.ni_max,
            v.mo_min,
            v.mo_max,
            v.mn_min,
            v.mn_max,
            v.cost_cr,
            v.cost_ni,
            v.cost_mo,
            v.cost_mn,
            v.sigma_req,
            v.hard_req,
            v.t_req,
            v.sum_min,
            v.sum_max,
            v.crni_max,
            coef["sigma_base"],
            coef["sigma_Cr"],
            coef["sigma_Mo"],
            coef["sigma_CrMo"],
            coef["hrc_base"],
            coef["hrc_Ni"],
            coef["hrc_Mn"],
            coef["hrc_NiMn"],
            coef["T_base"],
            coef["T_drop"],
            None,                       # created_at
            None,                       # note
        ]

        for col_idx, value in enumerate(row, start=1):
            ws_variants.cell(row=current_row, column=col_idx, value=value)

        current_row += 1

    # ---- ЛИСТ РЕЗУЛЬТАТЫ ----
    current_row = start_row

    for idx, r in enumerate(results, start=1):
        if current_row > start_row:
            copy_row_style(ws_results, start_row, current_row, max_col_results)

        req = result_req(r)
        limits = result_limits(r)

        if r.variant_id and r.variant_id in variant_key_map:
            excel_key = variant_key_map[r.variant_id]
            variant_name = r.variant.name if r.variant else "—"
            source_variant_id = r.variant_id
        else:
            excel_key = f"AUTO_R_{r.id}"
            variant_name = f"Из результата {r.id}"
            source_variant_id = None

        cr = to_float(r.cr)
        ni = to_float(r.ni)
        mo = to_float(r.mo)
        mn = to_float(r.mn)

        sum_additives = cr + ni + mo + mn
        cr_ni_product = cr * ni

        meets_sigma = ge_with_tol(r.sigma, req.get("sigma", 0))
        meets_hardness = ge_with_tol(r.hardness, req.get("hrc", 0))
        meets_t_melt = ge_with_tol(r.t_melt, req.get("t", 0))

        meets_sum_limit = (
                ge_with_tol(sum_additives, limits.get("sum_min", 0))
                and le_with_tol(sum_additives, limits.get("sum_max", 6))
        )

        meets_crni_limit = le_with_tol(cr_ni_product, limits.get("crni_max", 2.0))

        is_feasible = all([
            meets_sigma,
            meets_hardness,
            meets_t_melt,
            meets_sum_limit,
            meets_crni_limit,
        ])

        row = [
            r.id,
            excel_key,
            variant_name,
            source_variant_id,
            r.cr,
            r.ni,
            r.mo,
            r.mn,
            sum_additives,
            r.sigma,
            r.hardness,
            r.t_melt,
            r.cost,
            req.get("sigma"),
            req.get("hrc"),
            req.get("t"),
            limits.get("sum_min"),
            limits.get("sum_max"),
            limits.get("crni_max"),
            cr_ni_product,
            "Да" if meets_sigma else "Нет",
            "Да" if meets_hardness else "Нет",
            "Да" if meets_t_melt else "Нет",
            "Да" if meets_sum_limit else "Нет",
            "Да" if meets_crni_limit else "Нет",
            "Да" if is_feasible else "Нет",
            r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else None,
            None,
        ]

        for col_idx, value in enumerate(row, start=1):
            ws_results.cell(row=current_row, column=col_idx, value=value)

        current_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output