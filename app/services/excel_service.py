from __future__ import annotations

from io import BytesIO
from typing import Any
from openpyxl import load_workbook

from .optimizer import default_params


VARIANTS_SHEET_NAME = "Варианты"


def _normalize_value(value: Any):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        return value
    return value


def _to_float(value: Any, default=None):
    value = _normalize_value(value)
    if value is None:
        return default
    if isinstance(value, str):
        value = value.replace(",", ".")
    return float(value)


def _to_str(value: Any, default=""):
    value = _normalize_value(value)
    if value is None:
        return default
    return str(value).strip()


def read_variants_from_excel(file_storage) -> list[dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)

    if VARIANTS_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'В файле отсутствует лист "{VARIANTS_SHEET_NAME}"')

    ws = wb[VARIANTS_SHEET_NAME]

    # 2 строка = технические ключи
    headers = [cell.value for cell in ws[2]]
    headers = [_to_str(h, "") for h in headers]

    if not headers or all(not h for h in headers):
        raise ValueError("Во 2 строке листа 'Варианты' не найдены технические имена полей")

    required_headers = {
        "name",
        "cr_min", "cr_max",
        "ni_min", "ni_max",
        "mo_min", "mo_max",
        "mn_min", "mn_max",
        "cost_cr", "cost_ni", "cost_mo", "cost_mn",
        "sigma_req", "hard_req", "t_req",
    }

    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise ValueError(f"В шаблоне отсутствуют обязательные колонки: {', '.join(missing)}")

    defaults = default_params()

    variants = []

    # данные начинаются с 3 строки
    for row in ws.iter_rows(min_row=3, values_only=True):
        row_data = dict(zip(headers, row))

        # если строка полностью пустая — пропускаем
        if all(_normalize_value(v) is None for v in row_data.values()):
            continue

        name = _to_str(row_data.get("name"))
        if not name:
            raise ValueError("У одного из вариантов не заполнено поле 'Название варианта'")

        coef = {
            "sigma_base": _to_float(row_data.get("sigma_base"), defaults["coef"]["sigma_base"]),
            "sigma_Cr": _to_float(row_data.get("sigma_Cr"), defaults["coef"]["sigma_Cr"]),
            "sigma_Mo": _to_float(row_data.get("sigma_Mo"), defaults["coef"]["sigma_Mo"]),
            "sigma_CrMo": _to_float(row_data.get("sigma_CrMo"), defaults["coef"]["sigma_CrMo"]),
            "hrc_base": _to_float(row_data.get("hrc_base"), defaults["coef"]["hrc_base"]),
            "hrc_Ni": _to_float(row_data.get("hrc_Ni"), defaults["coef"]["hrc_Ni"]),
            "hrc_Mn": _to_float(row_data.get("hrc_Mn"), defaults["coef"]["hrc_Mn"]),
            "hrc_NiMn": _to_float(row_data.get("hrc_NiMn"), defaults["coef"]["hrc_NiMn"]),
            "T_base": _to_float(row_data.get("T_base"), defaults["coef"]["T_base"]),
            "T_drop": _to_float(row_data.get("T_drop"), defaults["coef"]["T_drop"]),
        }

        variant_data = {
            "name": name,
            "cr_min": _to_float(row_data.get("cr_min")),
            "cr_max": _to_float(row_data.get("cr_max")),
            "ni_min": _to_float(row_data.get("ni_min")),
            "ni_max": _to_float(row_data.get("ni_max")),
            "mo_min": _to_float(row_data.get("mo_min")),
            "mo_max": _to_float(row_data.get("mo_max")),
            "mn_min": _to_float(row_data.get("mn_min")),
            "mn_max": _to_float(row_data.get("mn_max")),
            "cost_cr": _to_float(row_data.get("cost_cr")),
            "cost_ni": _to_float(row_data.get("cost_ni")),
            "cost_mo": _to_float(row_data.get("cost_mo")),
            "cost_mn": _to_float(row_data.get("cost_mn")),
            "sigma_req": _to_float(row_data.get("sigma_req")),
            "hard_req": _to_float(row_data.get("hard_req")),
            "t_req": _to_float(row_data.get("t_req")),
            "sum_min": _to_float(row_data.get("sum_min"), defaults["limits"]["sum_min"]),
            "sum_max": _to_float(row_data.get("sum_max"), defaults["limits"]["sum_max"]),
            "crni_max": _to_float(row_data.get("crni_max"), defaults["limits"]["crni_max"]),
            "coef": coef,
        }

        # базовая валидация
        for field in [
            "cr_min", "cr_max", "ni_min", "ni_max", "mo_min", "mo_max", "mn_min", "mn_max",
            "cost_cr", "cost_ni", "cost_mo", "cost_mn",
            "sigma_req", "hard_req", "t_req"
        ]:
            if variant_data[field] is None:
                raise ValueError(f"У варианта '{name}' не заполнено обязательное поле '{field}'")

        for pair in [("cr_min", "cr_max"), ("ni_min", "ni_max"), ("mo_min", "mo_max"), ("mn_min", "mn_max")]:
            if variant_data[pair[0]] > variant_data[pair[1]]:
                raise ValueError(
                    f"У варианта '{name}' поле '{pair[0]}' больше '{pair[1]}'"
                )

        variants.append(variant_data)

    if not variants:
        raise ValueError("В файле не найдено ни одного варианта для загрузки")

    return variants